import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from utils.logger import setup_logger

logger = setup_logger()


def write_data_to_excel(
    products_data: dict[str, dict[str, str | None]],
    filename: str = "products.xlsx",
) -> None:
    """Записывает данные о продуктах в Excel-файл."""
    if not products_data:
        return

    df = pd.DataFrame.from_dict(products_data, orient="index")
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Products", index=False)
        worksheet = writer.sheets["Products"]

        for col_idx, column_cells in enumerate(worksheet.iter_cols(), start=1):
            max_length = max(
                (len(str(cell.value)) for cell in column_cells if cell.value), default=0
            )
            worksheet.column_dimensions[get_column_letter(col_idx)].width = (
                max_length + 2
            )

        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
