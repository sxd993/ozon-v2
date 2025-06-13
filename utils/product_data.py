import asyncio
import pandas as pd
from typing import Optional, Tuple
from bs4 import BeautifulSoup, Tag
from playwright.async_api import Page
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from utils.logger import setup_logger
import gc
import os

logger = setup_logger()


async def _get_stars_reviews(
    soup: BeautifulSoup,
) -> Tuple[Optional[str], Optional[str]]:
    """Извлекает рейтинг и количество отзывов."""
    try:
        product_statistic = soup.select_one("div[data-widget='webSingleProductScore']")
        if product_statistic and " • " in product_statistic.text:
            stars, reviews = product_statistic.text.strip().split(" • ")
            return stars.strip(), reviews.strip()
        return None, None
    except Exception as e:
        logger.warning(f"Ошибка при извлечении рейтинга/отзывов: {e}")
        return None, None


async def _get_sale_price(soup: BeautifulSoup) -> Optional[str]:
    """Извлекает цену с Ozon Картой."""
    try:
        price_element = soup.find(
            "span", string=lambda text: text and "Ozon Карт" in text
        )
        if price_element and price_element.parent:
            price_span = price_element.parent.select_one("div > span")
            if price_span:
                return (
                    price_span.text.strip()
                    .replace("\u2009", "")
                    .replace("₽", "")
                    .strip()
                )
        return None
    except Exception as e:
        logger.warning(f"Ошибка при извлечении цены с Ozon Картой: {e}")
        return None


async def _get_full_prices(soup: BeautifulSoup) -> Tuple[Optional[str], Optional[str]]:
    """Извлекает цену до скидок и без Ozon Карты."""
    try:
        price_element = soup.find(
            "span", string=lambda text: text and "без Ozon Карты" in text
        )
        if price_element and price_element.parent and price_element.parent.parent:
            price_spans = price_element.parent.parent.select("div > span")
            if price_spans:
                discount_price = (
                    price_spans[0]
                    .text.strip()
                    .replace("\u2009", "")
                    .replace("₽", "")
                    .strip()
                )
                base_price = (
                    price_spans[1]
                    .text.strip()
                    .replace("\u2009", "")
                    .replace("₽", "")
                    .strip()
                    if len(price_spans) > 1
                    else None
                )
                return discount_price, base_price
        return None, None
    except Exception as e:
        logger.warning(f"Ошибка при извлечении полной цены: {e}")
        return None, None


async def _get_product_name(soup: BeautifulSoup) -> str:
    """Извлекает название товара."""
    try:
        heading_div = soup.select_one("div[data-widget='webProductHeading']")
        if isinstance(heading_div, Tag):
            title_element = heading_div.find("h1")
            if isinstance(title_element, Tag):
                return title_element.text.strip().replace("\t", "").replace("\n", " ")
        return ""
    except Exception as e:
        logger.warning(f"Ошибка при извлечении названия товара: {e}")
        return ""


async def _get_salesman_name(soup: BeautifulSoup) -> Optional[str]:
    """Извлекает имя продавца."""
    try:
        for element in soup.select("a[href*='/seller/']"):
            href = element.get("href", "").lower()
            text = element.text.strip()
            if "reviews" not in href and "info" not in href and len(text) >= 2:
                return text
        return None
    except Exception as e:
        logger.warning(f"Ошибка при извлечении имени продавца: {e}")
        return None


async def _get_product_id(page: Page) -> Optional[str]:
    """Извлекает артикул товара."""
    try:
        element = await page.wait_for_selector(
            '//div[contains(text(), "Артикул: ")]', timeout=5000, state="attached"
        )
        text = await element.inner_text()
        return text.split("Артикул: ")[1].strip()
    except Exception as e:
        logger.warning(f"Ошибка при извлечении артикула: {e}")
        return None


async def _get_product_brand(soup: BeautifulSoup) -> Optional[str]:
    """Извлекает бренд товара."""
    try:
        breadcrumbs = soup.select_one("div[data-widget='breadCrumbs']")
        if breadcrumbs:
            last_item = breadcrumbs.select_one("li:last-child")
            if last_item:
                brand_tag = last_item.find("span")
                if brand_tag:
                    return brand_tag.get_text(strip=True)
        return None
    except Exception as e:
        logger.warning(f"Ошибка при извлечении бренда: {e}")
        return None


async def get_ozon_seller_info(page: Page) -> Tuple[Optional[str], Optional[str]]:
    """Извлекает информацию о продавце и ИНН."""
    try:
        seller_block = await page.wait_for_selector(
            "div[data-widget='webCurrentSeller']", timeout=5000, state="attached"
        )
        button = await seller_block.query_selector(
            "button:has(svg path[d='M8 0c4.964 0 8 3.036 8 8s-3.036 8-8 8-8-3.036-8-8 3.036-8 8-8m-.889 11.556a.889.889 0 0 0 1.778 0V8A.889.889 0 0 0 7.11 8zM8.89 4.444a.889.889 0 1 0-1.778 0 .889.889 0 0 0 1.778 0'])"
        )
        if not button:
            return None, None
        try:
            await button.click()
        except Exception:
            await page.evaluate("button => button.click()", button)
        await asyncio.sleep(0.3)

        modal = await page.wait_for_selector(
            "div[data-popper-placement^='top']", timeout=5000, state="visible"
        )
        if not modal:
            return None, None
        soup = BeautifulSoup(await page.content(), "lxml")
        modal_div = soup.select_one("div[data-popper-placement^='top']")
        if not modal_div:
            return None, None

        paragraphs = modal_div.find_all("p")
        seller_details = paragraphs[0].get_text(strip=True) if paragraphs else None
        inn = paragraphs[1].get_text(strip=True) if len(paragraphs) > 1 else None
        return seller_details, inn
    except Exception as e:
        logger.warning(f"Ошибка при извлечении информации о продавце: {e}")
        return None, None
    finally:
        if "soup" in locals():
            soup.decompose()
        gc.collect()


async def collect_product_info(
    page: Page, url: str, max_retries: int = 3
) -> dict[str, Optional[str]]:
    """Собирает информацию о товаре с сайта Ozon с повторными попытками."""
    for attempt in range(max_retries):
        try:
            logger.info(f"Попытка {attempt + 1}/{max_retries} обработки {url}")
            await page.goto(url, wait_until="domcontentloaded", timeout=30000)
            await asyncio.sleep(0.2)
            await page.wait_for_selector(
                "div[data-widget='webProductHeading']", timeout=5000, state="attached"
            )
            soup = BeautifulSoup(await page.content(), "lxml")

            seller_href = None
            try:
                seller_block = await page.wait_for_selector(
                    "div[data-widget='webCurrentSeller']",
                    timeout=5000,
                    state="attached",
                )
                seller_link = await seller_block.query_selector("a[href]")
                seller_href = (
                    await seller_link.get_attribute("href") if seller_link else None
                )
            except Exception as e:
                logger.warning(f"Ошибка при получении ссылки на продавца: {e}")

            product_id = await _get_product_id(page)
            product_name = await _get_product_name(soup)
            product_stars, product_reviews = await _get_stars_reviews(soup)
            product_ozon_card_price = await _get_sale_price(soup)
            product_discount_price, product_base_price = await _get_full_prices(soup)
            salesman = await _get_salesman_name(soup)
            product_brand = await _get_product_brand(soup)
            seller_details, inn = await get_ozon_seller_info(page)

            logger.info(f"Успешно собраны данные для {url}")
            return {
                "Артикул": product_id,
                "Название товара": product_name,
                "Бренд": product_brand,
                "Цена с картой озона": product_ozon_card_price,
                "Цена со скидкой": product_discount_price,
                "Цена": product_base_price,
                "Рейтинг": product_stars,
                "Отзывы": product_reviews,
                "Продавец": salesman,
                "Ссылка на продавца": seller_href,
                "Данные о продавце": seller_details,
                "ИНН": inn,
                "Ссылка на товар": url,
            }
        except Exception as e:
            logger.warning(f"Ошибка при обработке {url} (попытка {attempt + 1}): {e}")
            if attempt < max_retries - 1:
                await asyncio.sleep(5)
            else:
                logger.error(f"Не удалось обработать {url} после {max_retries} попыток")
                return {
                    "Артикул": None,
                    "Название товара": None,
                    "Бренд": None,
                    "Цена с картой озона": None,
                    "Цена со скидкой": None,
                    "Цена": None,
                    "Рейтинг": None,
                    "Отзывы": None,
                    "Продавец": None,
                    "Ссылка на продавца": None,
                    "Данные о продавце": None,
                    "ИНН": None,
                    "Ссылка на товар": url,
                }
        finally:
            if "soup" in locals():
                soup.decompose()
            gc.collect()


async def collect_data(
    products_urls: dict[str, str],
    page: Page,
    progress_handler=None,
    output_file: str = "ozon_products.xlsx",
    processed_file: str = "processed_links.txt",
) -> None:
    """Асинхронная функция сбора данных с использованием Playwright."""
    products_data = {}
    if progress_handler:
        progress_handler.set_total(len(products_urls))
    processed_count = 0

    for url in products_urls.values():
        processed_count += 1
        logger.info(f"Обработка товара {processed_count}/{len(products_urls)}: {url}")
        data = await collect_product_info(page=page, url=url)
        product_id = data.get("Артикул")
        if product_id and product_id not in products_data:
            products_data[product_id] = data
            # Сохраняем URL в файл обработанных ссылок
            try:
                with open(processed_file, "a", encoding="utf-8") as f:
                    f.write(f"{url}\n")
                logger.debug(f"URL {url} добавлен в {processed_file}")
            except Exception as e:
                logger.warning(f"Ошибка при записи в {processed_file}: {e}")
        if progress_handler:
            progress_handler.update()

        if processed_count % 10 == 0:
            write_data_to_excel(products_data=products_data, filename=output_file)
            gc.collect()
            logger.debug("Промежуточная запись в Excel и очистка памяти")

    if products_data:
        write_data_to_excel(products_data=products_data, filename=output_file)
        gc.collect()
        logger.info(f"Финальные данные сохранены в {output_file}")


def write_data_to_excel(
    products_data: dict[str, dict[str, str | None]], filename: str = "products.xlsx"
) -> None:
    """Записывает данные о продуктах в Excel-файл."""
    if not products_data:
        logger.warning("Нет данных для записи в Excel")
        return

    logger.info(f"Запись данных в {filename}")
    df = pd.DataFrame.from_dict(products_data, orient="index")
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Products", index=False)
        worksheet = writer.sheets["Products"]

        for col_idx, column_cells in enumerate(worksheet.iter_cols(), start=1):
            max_length = max(
                (len(str(cell.value)) for cell in column_cells if cell.value), default=0
            )
            worksheet.column_dimensions[get_column_letter(col_idx)].width = (
                max_length + 2
            )

        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
    logger.info(f"Excel-файл {filename} успешно создан")
